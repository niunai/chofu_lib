import os
import sys
import glob
import openpyxl
import util_gspread
from os.path import join, dirname
from dotenv import load_dotenv

dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

LOCAL_EXCEL_FILE = os.environ.get("LOCAL_EXCEL_FILE")

files = glob.glob(r".\temp\RENT*.csv")

if not files:
    print("No RENT*.csv file !")
    sys.exit()

for file in files:
    path = file

csv = ""

with open(path, encoding="utf-8_sig") as f:
    next(f)
    next(f)
    next(f)
    next(f)
    for line in f:
        csv = csv + line.replace('"', "")

os.remove(path)

wb = openpyxl.load_workbook(LOCAL_EXCEL_FILE)
ws = wb["Sheet1"]

last10title = []
for row in range(0, 10):
    last10title.append(ws.cell(ws.max_row - row, 3).value)

books = []
i = ws.max_row
for row in csv.split("\n"):
    cols = row.split(",")

    # print(len(cols), cols)
    if len(cols) < 2:
        break

    book = []
    title = cols[2]
    if title not in last10title:
        i += 1
        j = 0
        for c in cols:
            book.append(c)
            j += 1
            if j == 1:
                ws.cell(i, j).value = int(c)
            else:
                ws.cell(i, j).value = c
        books.append(book)

wb.save(LOCAL_EXCEL_FILE)

util_gspread.gspreadsheet_append(books)
